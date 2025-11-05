import pandas as pd
from openpyxl import load_workbook

# --- Step 1: Load your lounge eligibility lookup ---
lookup = pd.read_excel('lounge_eligibility_lookup_table.xlsx')

# Rename columns to match submission template
lookup = lookup.rename(columns={
    'tier1 %':'tier1 1%',
    'tier2 %':'tier2 1%',
    'tier3 %':'tier3 1%'
})

# Add Notes column
lookup['Notes'] = "Median percentage based on historical data; can be adjusted for future schedules"

# Reorder columns to match template
lookup = lookup[['grouping','example_destination','tier1 1%','tier2 1%','tier3 1%','Notes']]

# --- Step 2: Load your Excel template ---
template_path = '/Users/beast/Desktop/British Airways Lounge Analysis/Lounge Eligibility Lookup Template.xlsx'
wb = load_workbook(template_path)

# --- Step 3: Remove old sheet if it exists ---
sheet_name = 'Lounge Eligibility Lookup Table'
if sheet_name in wb.sheetnames:
    std = wb[sheet_name]
    wb.remove(std)
    wb.save(template_path)

# --- Step 4: Write the new table ---
with pd.ExcelWriter(template_path, engine='openpyxl', mode='a') as writer:
    lookup.to_excel(writer, sheet_name=sheet_name, index=False)

print("✅ Lounge Eligibility Table updated successfully!")
